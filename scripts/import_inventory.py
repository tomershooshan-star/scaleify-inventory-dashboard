"""
CSV/XLSX inventory importer.
Reads inventory files, normalizes columns, calculates days_left and tier,
then upserts into the SQLite database.
"""

import argparse
import csv
import json
import os
import sys

from dotenv import load_dotenv

load_dotenv()

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BASE_DIR)

from scripts.database import init_db, upsert_items_batch

# Column name mappings -- handles common variations from different ERP exports
COLUMN_MAP = {
    # SKU
    "sku": "sku",
    "item_code": "sku",
    "item code": "sku",
    "product_code": "sku",
    "product code": "sku",
    "part_number": "sku",
    "part number": "sku",
    "item_id": "sku",
    # Product name
    "product": "product",
    "product_name": "product",
    "product name": "product",
    "item_name": "product",
    "item name": "product",
    "description": "product",
    "item_description": "product",
    "name": "product",
    # Category
    "category": "category",
    "product_category": "category",
    "product category": "category",
    "group": "category",
    "item_group": "category",
    "type": "category",
    # Location
    "location": "location",
    "warehouse": "location",
    "warehouse_name": "location",
    "warehouse name": "location",
    "site": "location",
    "facility": "location",
    "store": "location",
    # Current stock
    "current_stock": "current_stock",
    "current stock": "current_stock",
    "stock": "current_stock",
    "qty": "current_stock",
    "quantity": "current_stock",
    "on_hand": "current_stock",
    "on hand": "current_stock",
    "qty_on_hand": "current_stock",
    "available": "current_stock",
    "inventory": "current_stock",
    # Daily demand
    "daily_demand": "daily_demand",
    "daily demand": "daily_demand",
    "avg_daily_demand": "daily_demand",
    "demand": "daily_demand",
    "daily_sales": "daily_demand",
    "daily sales": "daily_demand",
    "velocity": "daily_demand",
    "sales_velocity": "daily_demand",
    # Unit cost
    "unit_cost": "unit_cost",
    "unit cost": "unit_cost",
    "cost": "unit_cost",
    "price": "unit_cost",
    "unit_price": "unit_cost",
    "cogs": "unit_cost",
    # Reorder quantity
    "reorder_qty": "reorder_qty",
    "reorder qty": "reorder_qty",
    "reorder_quantity": "reorder_qty",
    "order_qty": "reorder_qty",
    "moq": "reorder_qty",
    "min_order": "reorder_qty",
    # Lead time
    "lead_time_days": "lead_time_days",
    "lead time days": "lead_time_days",
    "lead_time": "lead_time_days",
    "lead time": "lead_time_days",
    "delivery_days": "lead_time_days",
    # Supplier
    "supplier": "supplier",
    "supplier_name": "supplier",
    "supplier name": "supplier",
    "vendor": "supplier",
    "vendor_name": "supplier",
}

REQUIRED_FIELDS = ["sku", "product", "current_stock"]


def load_alert_rules() -> dict:
    """Load alert tier thresholds from config."""
    rules_path = os.path.join(BASE_DIR, "config", "alert_rules.json")
    if os.path.exists(rules_path):
        with open(rules_path, "r", encoding="utf-8") as f:
            return json.load(f)
    # Defaults if config missing
    return {
        "tiers": {
            "critical": {"days_left_max": 3},
            "warning": {"days_left_max": 7},
            "watch": {"days_left_max": 14},
        }
    }


def calculate_tier(days_left: float, rules: dict) -> str:
    """Determine stock tier based on days_left and configured thresholds."""
    tiers = rules.get("tiers", {})

    if days_left <= 0:
        return "out"

    critical_max = tiers.get("critical", {}).get("days_left_max", 3)
    warning_max = tiers.get("warning", {}).get("days_left_max", 7)
    watch_max = tiers.get("watch", {}).get("days_left_max", 14)

    if days_left <= critical_max:
        return "critical"
    elif days_left <= warning_max:
        return "warning"
    elif days_left <= watch_max:
        return "watch"
    else:
        return "healthy"


def normalize_columns(headers: list) -> dict:
    """Map raw CSV headers to standard column names. Returns {index: standard_name}."""
    mapping = {}
    for i, header in enumerate(headers):
        normalized = header.strip().lower().replace("-", "_")
        if normalized in COLUMN_MAP:
            mapping[i] = COLUMN_MAP[normalized]
    return mapping


def parse_number(value: str, default=0) -> float:
    """Safely parse a numeric value from a string."""
    if not value:
        return default
    try:
        # Remove currency symbols, commas, spaces
        cleaned = value.strip().replace(",", "").replace("$", "").replace(" ", "")
        return float(cleaned)
    except (ValueError, TypeError):
        return default


def read_csv_file(file_path: str) -> list:
    """Read a CSV file and return rows as dicts with raw values."""
    rows = []
    # Try different encodings
    for encoding in ["utf-8", "utf-8-sig", "latin-1", "cp1252"]:
        try:
            with open(file_path, "r", encoding=encoding, newline="") as f:
                # Sniff delimiter
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                except csv.Error:
                    dialect = csv.excel

                reader = csv.reader(f, dialect)
                headers = next(reader)
                col_map = normalize_columns(headers)

                if not col_map:
                    print(f"  WARNING: Could not map any columns in {file_path}")
                    print(f"  Headers found: {headers}")
                    return []

                for row_num, row in enumerate(reader, start=2):
                    item = {}
                    for idx, std_name in col_map.items():
                        if idx < len(row):
                            item[std_name] = row[idx].strip()
                    rows.append((row_num, item))
            break
        except UnicodeDecodeError:
            continue

    return rows


def read_xlsx_file(file_path: str) -> list:
    """Read an XLSX file and return rows as dicts with raw values."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  ERROR: openpyxl required for XLSX files. Run: pip install openpyxl")
        return []

    rows = []
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=False)):
        headers.append(str(cell.value or "").strip())

    col_map = normalize_columns(headers)
    if not col_map:
        print(f"  WARNING: Could not map any columns in {file_path}")
        print(f"  Headers found: {headers}")
        wb.close()
        return []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        item = {}
        for idx, std_name in col_map.items():
            if idx < len(row):
                val = row[idx]
                item[std_name] = str(val).strip() if val is not None else ""
        rows.append((row_num, item))

    wb.close()
    return rows


def import_file(file_path: str, default_location: str = "") -> dict:
    """
    Import a single CSV or XLSX file into the database.
    Returns: {"new": count, "updated": count, "skipped": count, "total": count}
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".csv":
        raw_rows = read_csv_file(file_path)
    elif ext in (".xlsx", ".xls"):
        raw_rows = read_xlsx_file(file_path)
    else:
        print(f"  Unsupported file format: {ext}")
        return {"new": 0, "updated": 0, "skipped": 0, "total": 0}

    if not raw_rows:
        return {"new": 0, "updated": 0, "skipped": 0, "total": 0}

    rules = load_alert_rules()
    items_to_upsert = []
    skipped = 0

    for row_num, item in raw_rows:
        # Validate required fields
        if not item.get("sku") or not item.get("product"):
            skipped += 1
            continue

        # Parse numeric fields
        current_stock = parse_number(item.get("current_stock", "0"))
        daily_demand = parse_number(item.get("daily_demand", "0"))
        unit_cost = parse_number(item.get("unit_cost", "0"))
        reorder_qty = int(parse_number(item.get("reorder_qty", "0")))
        lead_time_days = int(parse_number(item.get("lead_time_days", "0")))

        # Calculate days left
        if daily_demand > 0:
            days_left = round(current_stock / daily_demand, 1)
        elif current_stock > 0:
            days_left = 999  # Stock exists but no demand data
        else:
            days_left = 0

        tier = calculate_tier(days_left, rules)
        location = item.get("location", default_location) or default_location

        items_to_upsert.append({
            "sku": item["sku"],
            "product": item["product"],
            "category": item.get("category", ""),
            "location": location,
            "current_stock": int(current_stock),
            "daily_demand": daily_demand,
            "days_left": days_left,
            "tier": tier,
            "supplier": item.get("supplier", ""),
            "unit_cost": unit_cost,
            "reorder_qty": reorder_qty,
            "lead_time_days": lead_time_days,
        })

    if items_to_upsert:
        upsert_items_batch(items_to_upsert)

    result = {
        "new": len(items_to_upsert),  # Simplified: can't distinguish new vs updated without pre-check
        "updated": 0,
        "skipped": skipped,
        "total": len(raw_rows),
    }

    return result


def main():
    parser = argparse.ArgumentParser(description="Import inventory data from CSV/XLSX")
    parser.add_argument("--file", "-f", help="Path to CSV or XLSX file to import")
    parser.add_argument("--folder", help="Path to folder containing import files")
    parser.add_argument("--location", "-l", default="", help="Default location if not in file")
    args = parser.parse_args()

    # Ensure database exists
    init_db()

    files_to_process = []

    if args.file:
        if not os.path.exists(args.file):
            print(f"ERROR: File not found: {args.file}")
            sys.exit(1)
        files_to_process.append(args.file)
    elif args.folder:
        folder = args.folder
        if not os.path.isdir(folder):
            print(f"ERROR: Folder not found: {folder}")
            sys.exit(1)
        for fname in os.listdir(folder):
            if fname.lower().endswith((".csv", ".xlsx")):
                files_to_process.append(os.path.join(folder, fname))
    else:
        # Default: look in data/imports/
        import_dir = os.path.join(BASE_DIR, "data", "imports")
        if os.path.isdir(import_dir):
            for fname in os.listdir(import_dir):
                if fname.lower().endswith((".csv", ".xlsx")):
                    files_to_process.append(os.path.join(import_dir, fname))

    if not files_to_process:
        print("No files found to import.")
        print(f"  Drop CSV/XLSX files in: {os.path.join(BASE_DIR, 'data', 'imports')}")
        print("  Or specify a file: python import_inventory.py --file path/to/file.csv")
        return

    total_imported = 0
    total_skipped = 0

    for file_path in files_to_process:
        print(f"Importing: {os.path.basename(file_path)}")
        result = import_file(file_path, default_location=args.location)
        total_imported += result["new"]
        total_skipped += result["skipped"]
        print(f"  -> {result['new']} items imported, {result['skipped']} skipped")

    print(f"\nImported {total_imported} items ({total_skipped} skipped) from {len(files_to_process)} file(s)")


if __name__ == "__main__":
    main()
